#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
JOB_PROGRESS_PREFIX = "JOB_PROGRESS::"
USS_METRIC_ROWS: tuple[tuple[str, str], ...] = (
    ("Toplam_Talep", "Toplam Talep"),
    ("Uretilen", "Uretilen Talep"),
    ("Stoktan_Karsilanan", "Stoktan Karsilanan Talep"),
    ("Uretimden_Karsilanan", "Uretimden Karsilanan Talep"),
    ("Karsilanan", "Karsilanan Talep"),
    ("Karsilanamayan", "Karsilanamayan Talep"),
    ("Stok", "Donem Sonu Stok"),
)
USS_METRIC_KEYS: tuple[str, ...] = tuple(metric_key for metric_key, _label in USS_METRIC_ROWS)
DEFAULT_COLOR_BUCKET = "RENKSIZ"


@dataclass
class MissingGramaj:
    malzeme: str
    adet: int = 0
    toplam_miktar: float = 0.0


def emit_progress(percent: int, phase_code: str, phase_label: str) -> None:
    payload = {
        "percent": max(0, min(100, int(round(percent)))),
        "phase_code": str(phase_code).strip(),
        "phase_label": str(phase_label).strip(),
    }
    print(f"{JOB_PROGRESS_PREFIX}{json.dumps(payload, ensure_ascii=True)}", flush=True)


def normalize_material(value: object) -> str:
    text = str(value).strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0+)?", text):
        return str(int(float(text)))
    return text


def expand_job_dates(start_raw: object, end_raw: object, day_count: int) -> list[date]:
    if day_count <= 0:
        return []
    try:
        start_dt = date.fromisoformat(str(start_raw).strip()[:10])
    except ValueError:
        return []
    try:
        end_dt = date.fromisoformat(str(end_raw).strip()[:10])
    except ValueError:
        end_dt = start_dt + timedelta(days=day_count - 1)

    dates: list[date] = []
    d = start_dt
    while d <= end_dt:
        dates.append(d)
        d += timedelta(days=1)

    # Fallback to day_count from start when source date range is inconsistent.
    if len(dates) != day_count:
        dates = [start_dt + timedelta(days=i) for i in range(day_count)]
    return dates


def sort_key_mixed(value: str) -> tuple[int, int | str]:
    if value.isdigit():
        return (0, int(value))
    return (1, value)


def parse_float_like(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")

    if "," in text and "." in text:
        # Keep last separator as decimal separator, remove the other as thousand separator.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        text = text.replace(".", "")

    try:
        return float(text)
    except ValueError:
        return None


def quote_sql_identifier(identifier: str) -> str:
    return '"' + str(identifier).replace('"', '""') + '"'


def normalize_lookup_token(value: object) -> str:
    text = str(value or "").strip().lower().replace("\u0131", "i")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.translate(
        str.maketrans(
            {
                "\u00ff": "s",
                "\u00fe": "s",
                "\u00f0": "g",
                "\u0153": "u",
                "\u00e5": "a",
                "\u00e4": "a",
                "\u00e3": "a",
                "\u00b1": "i",
                "\u00a7": "c",
            }
        )
    )
    return "".join(ch for ch in text if ch.isalnum())


def resolve_stock_columns(column_names: list[str]) -> tuple[str, str, str] | None:
    parsed = [(name, normalize_lookup_token(name)) for name in column_names]
    if not parsed:
        return None

    def find_column(predicate) -> str | None:
        for original, token in parsed:
            if predicate(token):
                return original
        return None

    factory_col = find_column(
        lambda token: token in {"fabrika", "locationcode", "uretimyeri", "uy"}
        or ("uretim" in token and "yeri" in token)
    )
    material_col = find_column(
        lambda token: token in {"malzeme", "itemcode", "malzemekodu"} or "malzeme" in token
    )
    qty_col = find_column(
        lambda token: token in {"toplamstokadet", "toplamstok", "stok", "qty", "quantity", "miktar"}
        or "stok" in token
        or "qty" in token
        or "quantity" in token
        or "miktar" in token
    )
    if factory_col is None or material_col is None or qty_col is None:
        return None
    return factory_col, material_col, qty_col


def parse_month_from_talepler_column(column_name: str) -> str | None:
    match = re.fullmatch(r"(0[1-9]|1[0-2])\.(\d{4})\s*ADT", str(column_name).strip(), flags=re.IGNORECASE)
    if not match:
        return None
    month_num, year = match.group(1), match.group(2)
    return f"{year}-{month_num}"


def collect_monthly_total_demand_from_talepler(etl_db: Path, unit: str) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)

    with sqlite3.connect(etl_db) as conn:
        conn.row_factory = sqlite3.Row
        table_exists = (
            conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='Talepler'"
            ).fetchone()
            is not None
        )
        if not table_exists:
            return {}

        columns = [str(row["name"]).strip() for row in conn.execute("PRAGMA table_info('Talepler')")]
        if not columns:
            return {}

        month_columns: list[tuple[str, str]] = []
        for column_name in columns:
            month = parse_month_from_talepler_column(column_name)
            if month:
                month_columns.append((column_name, month))
        if not month_columns:
            return {}

        net_weight_column = next((name for name in columns if name == "Net ağırlığı"), None)
        if unit == "ton" and not net_weight_column:
            return {}

        for row in conn.execute('SELECT * FROM "Talepler"'):
            net_weight = parse_float_like(row[net_weight_column]) if net_weight_column else None
            for column_name, month in month_columns:
                qty = parse_float_like(row[column_name])
                if qty is None:
                    continue
                scaled_qty = qty * 1000.0
                if unit == "adet":
                    totals[month] += scaled_qty
                    continue
                if net_weight is None or net_weight <= 0:
                    continue
                totals[month] += scaled_qty * net_weight / 1_000_000.0

    return dict(totals)


def load_gramaj_map(etl_db: Path) -> dict[str, float]:
    gramaj_by_material: dict[str, float] = {}
    with sqlite3.connect(etl_db) as conn:
        conn.row_factory = sqlite3.Row
        
        # 1. Try to read from original 'Talepler' table ("Net ağırlığı" column)
        talepler_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='Talepler'"
        ).fetchone()
        
        if talepler_exists:
            cols = [str(r["name"]) for r in conn.execute("PRAGMA table_info('Talepler')")]
            if "Net ağırlığı" in cols and "Malzeme" in cols:
                cur = conn.execute('SELECT "Malzeme", "Net ağırlığı" FROM "Talepler"')
                for row in cur:
                    material = normalize_material(row["Malzeme"])
                    weight = parse_float_like(row["Net ağırlığı"])
                    if material and weight and weight > 0:
                        gramaj_by_material.setdefault(material, weight)

        # 2. Fallback to 'talepler_x' for any missing materials
        talepler_x_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='talepler_x'"
        ).fetchone()
        
        if talepler_x_exists:
            cur = conn.execute('SELECT "Malzeme", "Gramaj" FROM "talepler_x" WHERE "Gramaj" IS NOT NULL AND "Gramaj" > 0')
            for row in cur:
                material = normalize_material(row["Malzeme"])
                if material and material not in gramaj_by_material:
                    gramaj_by_material[material] = float(row["Gramaj"])
                    
    return gramaj_by_material


def load_fiili_gramaj_sum_map(etl_db: Path) -> dict[str, float]:
    """Malzeme_Data -> Parametre join'i uzerinden malzeme bazli tekil Fiili Gramaj katsayisi (SUM degil)."""
    gramaj_sum_by_material: dict[str, float] = {}
    with sqlite3.connect(etl_db) as conn:
        conn.row_factory = sqlite3.Row
        table_names = {"Malzeme_Data", "Parametre"}
        for table_name in table_names:
            table_exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
                (table_name,),
            ).fetchone()
            if table_exists is None:
                return {}

        md_columns = [str(row["name"]) for row in conn.execute("PRAGMA table_info('Malzeme_Data')")]
        p_columns = [str(row["name"]) for row in conn.execute("PRAGMA table_info('Parametre')")]
        if not md_columns or not p_columns:
            return {}

        def find_col(columns: list[str], normalized_name: str) -> str | None:
            for col in columns:
                if normalize_lookup_token(col) == normalized_name:
                    return col
            return None

        md_material_col = find_col(md_columns, "malzeme")
        md_mold_col = find_col(md_columns, "kalipcamambalaj")
        p_mold_col = find_col(p_columns, "kalip")
        p_gram_col = find_col(p_columns, "fiiligramaj")
        if not md_material_col or not md_mold_col or not p_mold_col or not p_gram_col:
            return {}

        query = f"""
            SELECT
                CAST(md.{quote_sql_identifier(md_material_col)} AS TEXT) AS material,
                COALESCE(MAX(p.{quote_sql_identifier(p_gram_col)}), 0.0) AS gramaj_sum
            FROM {quote_sql_identifier('Malzeme_Data')} md
            LEFT JOIN {quote_sql_identifier('Parametre')} p
                ON CAST(md.{quote_sql_identifier(md_mold_col)} AS TEXT) =
                   CAST(p.{quote_sql_identifier(p_mold_col)} AS TEXT)
            GROUP BY CAST(md.{quote_sql_identifier(md_material_col)} AS TEXT)
        """
        for row in conn.execute(query):
            material = normalize_material(row["material"])
            if not material:
                continue
            gramaj_sum = parse_float_like(row["gramaj_sum"])
            if gramaj_sum is None or gramaj_sum <= 0:
                continue
            gramaj_sum_by_material[material] = float(gramaj_sum)
    return gramaj_sum_by_material


def resolve_production_ton_gramaj_map(etl_db: Path) -> dict[str, float]:
    """
    Ton bazli uretim metriklerinde tek gramaj politikasi:
    once Fiili Gramaj, yoksa talepler gramaj fallback.
    """
    gramaj_by_material = load_fiili_gramaj_sum_map(etl_db)
    if gramaj_by_material:
        return gramaj_by_material
    return load_gramaj_map(etl_db)


def convert_quantity_by_unit(
    material: str,
    quantity: float,
    unit: str,
    gramaj_by_material: dict[str, float],
) -> float | None:
    if unit == "adet":
        return quantity
    gramaj = gramaj_by_material.get(material)
    if gramaj is None or gramaj <= 0:
        return None
    return quantity * gramaj / 1_000_000.0


def collect_monthly_tons_by_gramaj(
    plan_db: Path, plan_table: str, gramaj_by_material: dict[str, float]
) -> tuple[dict[tuple[str, str, str, str], float], list[str], dict[str, MissingGramaj]]:
    tons: dict[tuple[str, str, str, str], float] = defaultdict(float)
    months: set[str] = set()
    missing: dict[str, MissingGramaj] = {}

    query = f"""
        SELECT
            Malzeme,
            Fabrika,
            Firin,
            Hat,
            Baslangic_Tarihi,
            Bitis_Tarihi,
            Uretim_Gun_Sayisi,
            Ilk_Gun_Uretim_Miktari,
            Diger_Gunler_Gunluk_Uretim_Miktari
        FROM "{plan_table}"
        WHERE Baslangic_Tarihi IS NOT NULL
          AND Uretim_Gun_Sayisi IS NOT NULL
    """
    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(query):
            line_key = (
                str(row["Fabrika"]).strip(),
                str(row["Firin"]).strip(),
                str(row["Hat"]).strip(),
            )
            if not all(line_key):
                continue
            day_count = int(row["Uretim_Gun_Sayisi"] or 0)
            dates = expand_job_dates(row["Baslangic_Tarihi"], row["Bitis_Tarihi"], day_count)
            if not dates:
                continue
            first_qty = float(row["Ilk_Gun_Uretim_Miktari"] or 0.0)
            other_qty = float(row["Diger_Gunler_Gunluk_Uretim_Miktari"] or 0.0)

            material = normalize_material(row["Malzeme"])
            gramaj = gramaj_by_material.get(material)
            if gramaj is None or gramaj <= 0:
                miss = missing.setdefault(material, MissingGramaj(malzeme=material))
                miss.adet += 1
                miss.toplam_miktar += first_qty + max(0, len(dates) - 1) * other_qty
                continue

            for idx, work_day in enumerate(dates):
                month = work_day.strftime("%Y-%m")
                months.add(month)
                qty = first_qty if idx == 0 else other_qty
                ton = qty * gramaj / 1_000_000.0
                key = (line_key[0], line_key[1], line_key[2], month)
                tons[key] += ton

    return tons, sorted(months), missing


def collect_monthly_adet_by_production(
    plan_db: Path, plan_table: str
) -> tuple[dict[tuple[str, str, str, str], float], list[str]]:
    amounts: dict[tuple[str, str, str, str], float] = defaultdict(float)
    months: set[str] = set()

    query = f"""
        SELECT
            Fabrika,
            Firin,
            Hat,
            Baslangic_Tarihi,
            Bitis_Tarihi,
            Uretim_Gun_Sayisi,
            Ilk_Gun_Uretim_Miktari,
            Diger_Gunler_Gunluk_Uretim_Miktari,
            Toplam_Uretim_Miktari
        FROM "{plan_table}"
        WHERE Baslangic_Tarihi IS NOT NULL
          AND (Uretim_Gun_Sayisi IS NOT NULL OR Toplam_Uretim_Miktari IS NOT NULL)
    """
    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(query):
            line_key = (
                str(row["Fabrika"]).strip(),
                str(row["Firin"]).strip(),
                str(row["Hat"]).strip(),
            )
            if not all(line_key):
                continue
            day_count = int(row["Uretim_Gun_Sayisi"] or 0)
            dates = expand_job_dates(row["Baslangic_Tarihi"], row["Bitis_Tarihi"], day_count)
            first_qty = float(row["Ilk_Gun_Uretim_Miktari"] or 0.0)
            other_qty = float(row["Diger_Gunler_Gunluk_Uretim_Miktari"] or 0.0)

            if dates:
                for idx, work_day in enumerate(dates):
                    month = work_day.strftime("%Y-%m")
                    qty = first_qty if idx == 0 else other_qty
                    key = (line_key[0], line_key[1], line_key[2], month)
                    amounts[key] += qty
                    months.add(month)
                continue

            # Fallback for inconsistent legacy rows without day breakdown.
            month = str(row["Baslangic_Tarihi"]).strip()[:7]
            if not re.fullmatch(r"\d{4}-\d{2}", month):
                continue
            qty = float(row["Toplam_Uretim_Miktari"] or 0.0)
            key = (line_key[0], line_key[1], line_key[2], month)
            amounts[key] += qty
            months.add(month)

    return amounts, sorted(months)


def collect_monthly_tons_by_firin_cekisi(
    plan_db: Path, plan_table: str
) -> tuple[dict[tuple[str, str, str, str], float], list[str]]:
    tons: dict[tuple[str, str, str, str], float] = defaultdict(float)
    months: set[str] = set()

    query = f"""
        SELECT
            Fabrika,
            Firin,
            Hat,
            Baslangic_Tarihi,
            Bitis_Tarihi,
            Uretim_Gun_Sayisi,
            Toplam_Firin_Cekisi
        FROM "{plan_table}"
        WHERE Baslangic_Tarihi IS NOT NULL
          AND Uretim_Gun_Sayisi IS NOT NULL
          AND Toplam_Firin_Cekisi IS NOT NULL
    """
    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(query):
            line_key = (
                str(row["Fabrika"]).strip(),
                str(row["Firin"]).strip(),
                str(row["Hat"]).strip(),
            )
            if not all(line_key):
                continue
            day_count = int(row["Uretim_Gun_Sayisi"] or 0)
            dates = expand_job_dates(row["Baslangic_Tarihi"], row["Bitis_Tarihi"], day_count)
            if not dates:
                continue
            total_ton = float(row["Toplam_Firin_Cekisi"] or 0.0)
            daily_ton = total_ton / len(dates) if dates else 0.0

            for work_day in dates:
                month = work_day.strftime("%Y-%m")
                months.add(month)
                key = (line_key[0], line_key[1], line_key[2], month)
                tons[key] += daily_ton

    return tons, sorted(months)


def collect_monthly_change_counts(
    plan_db: Path, plan_table: str
) -> tuple[dict[tuple[str, str, str, str], float], list[str]]:
    # For each Fabrika-Firin-Hat and month, count material switches in job order.
    job_rows: dict[tuple[str, str, str, str], list[tuple[str, int, str]]] = defaultdict(list)
    months: set[str] = set()

    query = f"""
        SELECT Is_ID, Malzeme, Fabrika, Firin, Hat, Baslangic_Tarihi
        FROM "{plan_table}"
        WHERE Baslangic_Tarihi IS NOT NULL
        ORDER BY Baslangic_Tarihi, Is_ID
    """
    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(query):
            month = str(row["Baslangic_Tarihi"]).strip()[:7]
            if not re.fullmatch(r"\d{4}-\d{2}", month):
                continue
            months.add(month)
            key = (
                str(row["Fabrika"]).strip(),
                str(row["Firin"]).strip(),
                str(row["Hat"]).strip(),
                month,
            )
            job_rows[key].append(
                (
                    str(row["Baslangic_Tarihi"]).strip(),
                    int(row["Is_ID"] or 0),
                    normalize_material(row["Malzeme"]),
                )
            )

    # Detect whether a line has an ongoing job at month start from previous month.
    carryover_at_month_start: dict[tuple[str, str, str, str], bool] = {}
    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        carry_q = f"""
            SELECT 1
            FROM "{plan_table}"
            WHERE Fabrika = ?
              AND Firin = ?
              AND Hat = ?
              AND Baslangic_Tarihi IS NOT NULL
              AND Bitis_Tarihi IS NOT NULL
              AND date(Baslangic_Tarihi) < date(?)
              AND date(Bitis_Tarihi) >= date(?)
            LIMIT 1
        """
        for fabrika, firin, hat, month in job_rows:
            month_start = f"{month}-01"
            key = (fabrika, firin, hat, month)
            carryover_at_month_start[key] = (
                conn.execute(carry_q, (fabrika, firin, hat, month_start, month_start)).fetchone() is not None
            )

    changes: dict[tuple[str, str, str, str], float] = defaultdict(float)
    for key, rows in job_rows.items():
        rows.sort(key=lambda x: (x[0], x[1]))
        change_count = 0
        prev_material = None
        for _dt, _is_id, material in rows:
            if prev_material is None:
                prev_material = material
                continue
            if material != prev_material:
                change_count += 1
            prev_material = material

        # If month's first job starts exactly on day 1 and no previous-month carryover exists,
        # count that first startup as a change event.
        first_day = rows[0][0][:10]
        if first_day.endswith("-01") and not carryover_at_month_start.get(key, False):
            change_count += 1

        changes[key] = float(change_count)

    return changes, sorted(months)


def load_playground_structure(
    plan_db: Path,
) -> dict[str, dict[str, list[str]]]:
    structure: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute('SELECT Fabrika, Firin, Hatlar FROM "Playground_x"'):
            fabrika = str(row["Fabrika"]).strip()
            firin = str(row["Firin"]).strip()
            hat = str(row["Hatlar"]).strip()
            if not fabrika or not firin or not hat:
                continue
            if hat not in structure[fabrika][firin]:
                structure[fabrika][firin].append(hat)
    for fabrika in structure:
        for firin in structure[fabrika]:
            structure[fabrika][firin] = sorted(structure[fabrika][firin], key=sort_key_mixed)
    return structure


def normalize_color_name(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return DEFAULT_COLOR_BUCKET
    text = re.sub(r"\s+", " ", text)
    if text.upper() in {"BILINMEYEN", "UNKNOWN", "N/A", "NA", "-"}:
        return DEFAULT_COLOR_BUCKET
    return text


def load_material_color_info(
    etl_db: Path,
) -> tuple[dict[str, str], list[str]]:
    material_to_color: dict[str, str] = {}
    color_order: list[str] = []
    color_seen: set[str] = set()
    talep_to_firin: dict[str, str] = {}

    with sqlite3.connect(etl_db) as conn:
        conn.row_factory = sqlite3.Row

        try:
            for row in conn.execute(
                'SELECT Talepteki_Renk, Firindaki_Renk FROM "Talep_Renk Map"'
            ):
                talep = normalize_color_name(row["Talepteki_Renk"])
                firin = normalize_color_name(row["Firindaki_Renk"])
                talep_to_firin[talep.upper()] = firin
                if firin not in color_seen:
                    color_seen.add(firin)
                    color_order.append(firin)
        except sqlite3.OperationalError:
            pass

        for row in conn.execute('SELECT Malzeme, Renk FROM "talepler_x" ORDER BY rowid'):
            material = normalize_material(row["Malzeme"])
            if not material:
                continue
            demand_color = normalize_color_name(row["Renk"])
            firin_color = talep_to_firin.get(demand_color.upper(), demand_color)
            material_to_color.setdefault(material, firin_color)
            if firin_color not in color_seen:
                color_seen.add(firin_color)
                color_order.append(firin_color)

    return material_to_color, color_order


def collect_color_factory_month_metrics(
    plan_db: Path,
    etl_db: Path,
    plan_table: str,
    allocation_table: str = "Full_ORTools_Plan_Is_Talep_Esleme",
    unit: str = "ton",
) -> tuple[
    list[str],
    list[str],
    dict[tuple[str, str, str], float],
    dict[tuple[str, str, str], float],
    dict[tuple[str, str, str], float],
    dict[str, float],
]:
    if unit not in {"ton", "adet"}:
        raise ValueError("unit must be 'ton' or 'adet'.")
    if unit == "ton":
        gramaj_by_material = resolve_production_ton_gramaj_map(etl_db)
    else:
        gramaj_by_material = {}
    material_to_color, color_order = load_material_color_info(etl_db)

    production_by_factory_material_month: dict[tuple[str, str, str], float] = defaultdict(float)
    fixed_alloc_by_material_month_factory: dict[tuple[str, str, str], float] = defaultdict(float)
    null_alloc_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    start_stock_by_factory_material: dict[tuple[str, str], float] = defaultdict(float)
    factories: set[str] = set()
    months: set[str] = set()

    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row

        job_factory_by_id: dict[int, str] = {}
        for row in conn.execute(f'SELECT Is_ID, Fabrika FROM "{plan_table}"'):
            is_id = int(row["Is_ID"] or 0)
            if is_id > 0:
                job_factory_by_id[is_id] = str(row["Fabrika"]).strip()

        for row in conn.execute(
            f"""
            SELECT
                Malzeme,
                Fabrika,
                Baslangic_Tarihi,
                Bitis_Tarihi,
                Uretim_Gun_Sayisi,
                Ilk_Gun_Uretim_Miktari,
                Diger_Gunler_Gunluk_Uretim_Miktari,
                Toplam_Uretim_Miktari
            FROM "{plan_table}"
            WHERE Baslangic_Tarihi IS NOT NULL
              AND (Uretim_Gun_Sayisi IS NOT NULL OR Toplam_Uretim_Miktari IS NOT NULL)
            """
        ):
            material = normalize_material(row["Malzeme"])
            factory = str(row["Fabrika"]).strip()
            if not material or not factory:
                continue

            day_count = int(row["Uretim_Gun_Sayisi"] or 0)
            dates = expand_job_dates(row["Baslangic_Tarihi"], row["Bitis_Tarihi"], day_count)
            first_qty = float(row["Ilk_Gun_Uretim_Miktari"] or 0.0)
            other_qty = float(row["Diger_Gunler_Gunluk_Uretim_Miktari"] or 0.0)

            if dates:
                for idx, work_day in enumerate(dates):
                    month = work_day.strftime("%Y-%m")
                    qty = first_qty if idx == 0 else other_qty
                    amount = convert_quantity_by_unit(material, qty, unit, gramaj_by_material)
                    if amount is None:
                        continue
                    production_by_factory_material_month[(factory, material, month)] += amount
                    months.add(month)
                factories.add(factory)
                continue

            # Fallback for inconsistent legacy rows without day breakdown.
            month = str(row["Baslangic_Tarihi"]).strip()[:7]
            if not re.fullmatch(r"\d{4}-\d{2}", month):
                continue
            qty = float(row["Toplam_Uretim_Miktari"] or 0.0)
            amount = convert_quantity_by_unit(material, qty, unit, gramaj_by_material)
            if amount is None:
                continue
            production_by_factory_material_month[(factory, material, month)] += amount
            factories.add(factory)
            months.add(month)

        try:
            for row in conn.execute(
                f"""
                SELECT Is_ID, Malzeme, Talep_Donemi, Atanan_Miktar
                FROM "{allocation_table}"
                WHERE Talep_Donemi IS NOT NULL AND Atanan_Miktar IS NOT NULL
                """
            ):
                month = str(row["Talep_Donemi"]).strip()[:7]
                if not re.fullmatch(r"\d{4}-\d{2}", month):
                    continue
                material = normalize_material(row["Malzeme"])
                amount = convert_quantity_by_unit(
                    material,
                    float(row["Atanan_Miktar"] or 0.0),
                    unit,
                    gramaj_by_material,
                )
                if amount is None:
                    continue
                is_id = row["Is_ID"]
                if is_id is None:
                    null_alloc_by_material_month[(material, month)] += amount
                else:
                    factory = job_factory_by_id.get(int(is_id))
                    if factory:
                        fixed_alloc_by_material_month_factory[(material, month, factory)] += amount
                        factories.add(factory)
                months.add(month)
        except sqlite3.OperationalError:
            pass

        for row in conn.execute('SELECT Talep_Donemi FROM "Full_ORTools_Plan_Talep_Stok"'):
            month = str(row["Talep_Donemi"]).strip()[:7]
            if re.fullmatch(r"\d{4}-\d{2}", month):
                months.add(month)

    with sqlite3.connect(etl_db) as conn:
        conn.row_factory = sqlite3.Row
        stock_cols = [str(r["name"]) for r in conn.execute("PRAGMA table_info('Stok')")]
        resolved_stock_columns = resolve_stock_columns(stock_cols)
        if resolved_stock_columns is not None:
            prod_col, material_col, qty_col = resolved_stock_columns
            for row in conn.execute(
                f"SELECT "
                f"{quote_sql_identifier(prod_col)} AS Factory, "
                f"{quote_sql_identifier(material_col)} AS Material, "
                f"{quote_sql_identifier(qty_col)} AS Qty "
                f'FROM {quote_sql_identifier("Stok")}'
            ):
                material = normalize_material(row["Material"])
                factory = str(row["Factory"]).strip()
                qty = float(row["Qty"] or 0.0)
                amount = convert_quantity_by_unit(material, qty, unit, gramaj_by_material)
                if amount is None:
                    continue
                start_stock_by_factory_material[(factory, material)] += amount
                factories.add(factory)

    months_list = sorted(months)
    factories_list = sorted(factories, key=sort_key_mixed)

    prod_by_factory_color_month: dict[tuple[str, str, str], float] = defaultdict(float)
    served_by_factory_color_month: dict[tuple[str, str, str], float] = defaultdict(float)
    stock_end_by_factory_color_month: dict[tuple[str, str, str], float] = defaultdict(float)

    all_materials: set[str] = set()
    for _factory, material, _month in production_by_factory_material_month:
        all_materials.add(material)
    for material, _month, _factory in fixed_alloc_by_material_month_factory:
        all_materials.add(material)
    for material, _month in null_alloc_by_material_month:
        all_materials.add(material)
    for _factory, material in start_stock_by_factory_material:
        all_materials.add(material)

    diagnostics = {
        "null_alloc_total_ton": 0.0,
        "null_alloc_distributed_ton": 0.0,
        "null_alloc_undistributed_ton": 0.0,
        "fixed_alloc_shortfall_ton": 0.0,
    }

    for material in all_materials:
        color = material_to_color.get(material, DEFAULT_COLOR_BUCKET)
        if color not in color_order:
            color_order.append(color)
        inventory = {factory: start_stock_by_factory_material.get((factory, material), 0.0) for factory in factories_list}

        for month in months_list:
            available = {}
            for factory in factories_list:
                prod_ton = production_by_factory_material_month.get((factory, material, month), 0.0)
                available[factory] = inventory[factory] + prod_ton
                prod_by_factory_color_month[(factory, color, month)] += prod_ton

            served_fixed = {}
            for factory in factories_list:
                req = fixed_alloc_by_material_month_factory.get((material, month, factory), 0.0)
                eff = min(req, max(available[factory], 0.0))
                served_fixed[factory] = eff
                if req > eff:
                    diagnostics["fixed_alloc_shortfall_ton"] += (req - eff)

            remaining = {f: max(available[f] - served_fixed[f], 0.0) for f in factories_list}

            null_req = null_alloc_by_material_month.get((material, month), 0.0)
            diagnostics["null_alloc_total_ton"] += null_req
            alloc_from_null = {f: 0.0 for f in factories_list}
            sum_remaining = sum(remaining.values())
            distributable = min(null_req, sum_remaining)
            if distributable > 0 and sum_remaining > 0:
                for factory in factories_list:
                    alloc_from_null[factory] = distributable * (remaining[factory] / sum_remaining)
                diagnostics["null_alloc_distributed_ton"] += distributable
                diagnostics["null_alloc_undistributed_ton"] += max(0.0, null_req - distributable)
            else:
                diagnostics["null_alloc_undistributed_ton"] += null_req

            for factory in factories_list:
                served_total = served_fixed[factory] + alloc_from_null[factory]
                served_by_factory_color_month[(factory, color, month)] += served_total
                inventory[factory] = max(available[factory] - served_total, 0.0)
                stock_end_by_factory_color_month[(factory, color, month)] += inventory[factory]

    return (
        months_list,
        color_order,
        prod_by_factory_color_month,
        stock_end_by_factory_color_month,
        served_by_factory_color_month,
        diagnostics,
    )


def collect_unmet_factory_color_month_metrics(
    plan_db: Path,
    etl_db: Path,
    unit: str = "ton",
    demand_table: str = "Full_ORTools_Plan_Talep_Stok",
) -> dict[tuple[str, str, str], float]:
    if unit not in {"ton", "adet"}:
        raise ValueError("unit must be 'ton' or 'adet'.")

    gramaj_by_material = load_gramaj_map(etl_db)
    material_to_color, _ = load_material_color_info(etl_db)
    unmet_by_factory_color_month: dict[tuple[str, str, str], float] = defaultdict(float)

    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(
            f"""
            SELECT Malzeme, Talep_Donemi, Kalan_Acik_Talep
            FROM "{demand_table}"
            WHERE Talep_Donemi IS NOT NULL
              AND Kalan_Acik_Talep IS NOT NULL
              AND Kalan_Acik_Talep > 0
            """
        ):
            month = str(row["Talep_Donemi"]).strip()[:7]
            if not re.fullmatch(r"\d{4}-\d{2}", month):
                continue
            material = normalize_material(row["Malzeme"])
            if not material:
                continue
            amount = convert_quantity_by_unit(
                material,
                float(row["Kalan_Acik_Talep"] or 0.0),
                unit,
                gramaj_by_material,
            )
            if amount is None:
                continue
            color = material_to_color.get(material, DEFAULT_COLOR_BUCKET)
            unmet_by_factory_color_month[("BELIRSIZ", color, month)] += amount

    return unmet_by_factory_color_month


def collect_uss_month_metrics(
    plan_db: Path,
    etl_db: Path,
    unit: str = "ton",
    demand_table: str = "Full_ORTools_Plan_Talep_Stok",
    plan_table: str = "Full_ORTools_Plan_Uretim_Isleri",
    allocation_table: str = "Full_ORTools_Plan_Is_Talep_Esleme",
) -> tuple[list[str], dict[str, dict[str, float]]]:
    if unit not in {"ton", "adet"}:
        raise ValueError("unit must be 'ton' or 'adet'.")

    if unit == "ton":
        gramaj_by_material = resolve_production_ton_gramaj_map(etl_db)
    else:
        gramaj_by_material = {}
    metric_columns = (
        ("Toplam_Talep_Fallback", "Donem_Talebi"),
        ("Uretilen_Talep", "Bu_Donemde_Uretilen"),
        ("Karsilanan", "Karsilanan_Talep_Miktari"),
        ("Stoktan_Karsilanan_Direct", "Stoktan_Karsilanan_Miktar"),
        ("Uretimden_Karsilanan_Direct", "Uretimden_Karsilanan_Miktar"),
        ("Karsilanamayan", "Kalan_Acik_Talep"),
        ("Donem_Sonu_Stok_Direct", "Donem_Sonu_Stok"),
    )
    all_metric_names = list(USS_METRIC_KEYS)
    metric_month_values: dict[str, dict[str, float]] = {
        metric_name: defaultdict(float) for metric_name in all_metric_names
    }
    fallback_toplam_talep_month_values: dict[str, float] = defaultdict(float)
    fallback_produced_for_demand_month_values: dict[str, float] = defaultdict(float)
    produced_for_demand_month_values: dict[str, float] = defaultdict(float)
    stock_served_month_values: dict[str, float] = defaultdict(float)
    months: set[str] = set()
    allocation_table_available = False

    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        available_columns = {
            str(row["name"]).strip()
            for row in conn.execute(f'PRAGMA table_info("{demand_table}")')
        }
        if not available_columns:
            return [], {metric_name: {} for metric_name in all_metric_names}
        has_direct_stock_column = "Stoktan_Karsilanan_Miktar" in available_columns
        has_direct_production_column = "Uretimden_Karsilanan_Miktar" in available_columns

        select_parts = ["Malzeme", "Talep_Donemi"]
        for metric_name, source_column in metric_columns:
            if source_column in available_columns:
                select_parts.append(f'"{source_column}" AS "{metric_name}"')
            else:
                select_parts.append(f'0.0 AS "{metric_name}"')
        query = f"""
            SELECT {", ".join(select_parts)}
            FROM "{demand_table}"
            WHERE Talep_Donemi IS NOT NULL
        """

        for row in conn.execute(query):
            month = str(row["Talep_Donemi"]).strip()[:7]
            if not re.fullmatch(r"\d{4}-\d{2}", month):
                continue
            months.add(month)
            material = normalize_material(row["Malzeme"])
            for metric_name, _source_column in metric_columns:
                quantity = float(row[metric_name] or 0.0)
                amount = convert_quantity_by_unit(material, quantity, unit, gramaj_by_material)
                if amount is None:
                    continue
                if metric_name == "Toplam_Talep_Fallback":
                    fallback_toplam_talep_month_values[month] += amount
                    continue
                if metric_name == "Uretilen_Talep":
                    fallback_produced_for_demand_month_values[month] += amount
                elif metric_name == "Stoktan_Karsilanan_Direct":
                    metric_month_values["Stoktan_Karsilanan"][month] += amount
                elif metric_name == "Uretimden_Karsilanan_Direct":
                    metric_month_values["Uretimden_Karsilanan"][month] += amount
                elif metric_name == "Donem_Sonu_Stok_Direct":
                    # USS stok satiri yalnizca Full_ORTools_Plan_Talep_Stok.Donem_Sonu_Stok kaynagindan toplanir.
                    metric_month_values["Stok"][month] += amount
                else:
                    metric_month_values[metric_name][month] += amount

        allocation_table_available = (
            conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
                (allocation_table,),
            ).fetchone()
            is not None
        )
        if allocation_table_available:
            for row in conn.execute(
                f"""
                SELECT Is_ID, Malzeme, Talep_Donemi, Atanan_Miktar
                FROM "{allocation_table}"
                WHERE Talep_Donemi IS NOT NULL
                  AND Atanan_Miktar IS NOT NULL
                """
            ):
                month = str(row["Talep_Donemi"]).strip()[:7]
                if not re.fullmatch(r"\d{4}-\d{2}", month):
                    continue
                material = normalize_material(row["Malzeme"])
                if not material:
                    continue
                amount = convert_quantity_by_unit(
                    material,
                    float(row["Atanan_Miktar"] or 0.0),
                    unit,
                    gramaj_by_material,
                )
                if amount is None:
                    continue
                months.add(month)
                if row["Is_ID"] is None:
                    stock_served_month_values[month] += amount
                else:
                    produced_for_demand_month_values[month] += amount

    for month, amount in fallback_toplam_talep_month_values.items():
        months.add(month)
        metric_month_values["Toplam_Talep"][month] += amount

    produced_by_month: dict[str, float] = defaultdict(float)
    with sqlite3.connect(plan_db) as conn:
        has_plan_table = (
            conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
                (plan_table,),
            ).fetchone()
            is not None
        )

    if allocation_table_available:
        for month, amount in produced_for_demand_month_values.items():
            produced_by_month[month] += amount
    elif has_plan_table:
        if unit == "adet":
            produced_by_line_month, _months_production = collect_monthly_adet_by_production(plan_db, plan_table)
            for (_fabrika, _firin, _hat, month), amount in produced_by_line_month.items():
                produced_by_month[month] += amount
                months.add(month)
        else:
            produced_by_line_month_ton, _months_production_ton, _missing = collect_monthly_tons_by_gramaj(
                plan_db, plan_table, gramaj_by_material
            )
            for (_fabrika, _firin, _hat, month), amount in produced_by_line_month_ton.items():
                produced_by_month[month] += amount
                months.add(month)
    else:
        for month, amount in fallback_produced_for_demand_month_values.items():
            produced_by_month[month] += amount

    # Keep USS end-of-period stock aligned with color-factory summary stock.
    # This includes non-demand materials and uses the same allocation/redistribution logic.
    color_stock_months, _color_rows, _prod_by_factory_color_month, stock_by_factory_color_month, _served_by_factory_color_month, _color_diag = (
        collect_color_factory_month_metrics(
            plan_db,
            etl_db,
            plan_table,
            allocation_table=allocation_table,
            unit=unit,
        )
    )
    has_etl_stock_table = False
    try:
        with sqlite3.connect(etl_db) as conn:
            has_etl_stock_table = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='Stok'"
            ).fetchone() is not None
    except Exception:
        pass

    stock_by_month: dict[str, float] = defaultdict(float)
    for (_factory, _color, month), amount in stock_by_factory_color_month.items():
        stock_by_month[month] += amount
    for month in color_stock_months:
        months.add(month)
    if has_etl_stock_table:
        for month, amount in stock_by_month.items():
            metric_month_values["Stok"][month] = amount

    months_list = sorted(months)
    for month in months_list:
        metric_month_values["Uretilen"][month] = produced_by_month.get(month, 0.0)
        if has_direct_stock_column:
            metric_month_values["Stoktan_Karsilanan"][month] = metric_month_values[
                "Stoktan_Karsilanan"
            ].get(month, 0.0)
        elif allocation_table_available:
            metric_month_values["Stoktan_Karsilanan"][month] = stock_served_month_values.get(month, 0.0)
        else:
            served_value = metric_month_values["Karsilanan"].get(month, 0.0)
            produced_for_demand = fallback_produced_for_demand_month_values.get(month, 0.0)
            metric_month_values["Stoktan_Karsilanan"][month] = max(served_value - produced_for_demand, 0.0)

        if has_direct_production_column:
            metric_month_values["Uretimden_Karsilanan"][month] = metric_month_values[
                "Uretimden_Karsilanan"
            ].get(month, 0.0)
        elif allocation_table_available:
            metric_month_values["Uretimden_Karsilanan"][month] = produced_for_demand_month_values.get(month, 0.0)
        else:
            served_value = metric_month_values["Karsilanan"].get(month, 0.0)
            stock_served = metric_month_values["Stoktan_Karsilanan"].get(month, 0.0)
            metric_month_values["Uretimden_Karsilanan"][month] = max(served_value - stock_served, 0.0)
    return (
        months_list,
        {
            metric_name: {month: metric_month_values[metric_name].get(month, 0.0) for month in months_list}
            for metric_name in all_metric_names
        },
    )


def collect_uss_color_month_metrics(
    plan_db: Path,
    etl_db: Path,
    unit: str = "ton",
    demand_table: str = "Full_ORTools_Plan_Talep_Stok",
    plan_table: str = "Full_ORTools_Plan_Uretim_Isleri",
    allocation_table: str = "Full_ORTools_Plan_Is_Talep_Esleme",
) -> tuple[list[str], list[str], dict[str, dict[str, dict[str, float]]]]:
    if unit not in {"ton", "adet"}:
        raise ValueError("unit must be 'ton' or 'adet'.")

    if unit == "ton":
        gramaj_by_material = resolve_production_ton_gramaj_map(etl_db)
    else:
        gramaj_by_material = {}

    material_to_color, color_order = load_material_color_info(etl_db)
    metric_columns = (
        ("Toplam_Talep_Fallback", "Donem_Talebi"),
        ("Uretilen_Talep", "Bu_Donemde_Uretilen"),
        ("Karsilanan", "Karsilanan_Talep_Miktari"),
        ("Stoktan_Karsilanan_Direct", "Stoktan_Karsilanan_Miktar"),
        ("Uretimden_Karsilanan_Direct", "Uretimden_Karsilanan_Miktar"),
        ("Karsilanamayan", "Kalan_Acik_Talep"),
        ("Donem_Sonu_Stok_Direct", "Donem_Sonu_Stok"),
    )

    fallback_toplam_talep_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    fallback_produced_for_demand_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    produced_for_demand_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    stock_served_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    karsilanan_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    karsilanamayan_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    stok_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    direct_stoktan_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    direct_uretimden_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    months: set[str] = set()
    allocation_table_available = False
    has_direct_stock_column = False
    has_direct_production_column = False

    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        available_columns = {
            str(row["name"]).strip()
            for row in conn.execute(f'PRAGMA table_info("{demand_table}")')
        }
        if not available_columns:
            return [], [], {}

        has_direct_stock_column = "Stoktan_Karsilanan_Miktar" in available_columns
        has_direct_production_column = "Uretimden_Karsilanan_Miktar" in available_columns

        select_parts = ["Malzeme", "Talep_Donemi"]
        for metric_name, source_column in metric_columns:
            if source_column in available_columns:
                select_parts.append(f'"{source_column}" AS "{metric_name}"')
            else:
                select_parts.append(f'0.0 AS "{metric_name}"')

        query = f"""
            SELECT {", ".join(select_parts)}
            FROM "{demand_table}"
            WHERE Talep_Donemi IS NOT NULL
        """
        for row in conn.execute(query):
            month = str(row["Talep_Donemi"]).strip()[:7]
            if not re.fullmatch(r"\d{4}-\d{2}", month):
                continue
            material = normalize_material(row["Malzeme"])
            if not material:
                continue
            months.add(month)
            key = (material, month)

            for metric_name, _source_column in metric_columns:
                quantity = float(row[metric_name] or 0.0)
                amount = convert_quantity_by_unit(material, quantity, unit, gramaj_by_material)
                if amount is None:
                    continue

                if metric_name == "Toplam_Talep_Fallback":
                    fallback_toplam_talep_by_material_month[key] += amount
                elif metric_name == "Uretilen_Talep":
                    fallback_produced_for_demand_by_material_month[key] += amount
                elif metric_name == "Karsilanan":
                    karsilanan_by_material_month[key] += amount
                elif metric_name == "Stoktan_Karsilanan_Direct":
                    direct_stoktan_by_material_month[key] += amount
                elif metric_name == "Uretimden_Karsilanan_Direct":
                    direct_uretimden_by_material_month[key] += amount
                elif metric_name == "Karsilanamayan":
                    karsilanamayan_by_material_month[key] += amount
                elif metric_name == "Donem_Sonu_Stok_Direct":
                    stok_by_material_month[key] += amount

        allocation_table_available = (
            conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
                (allocation_table,),
            ).fetchone()
            is not None
        )
        if allocation_table_available:
            for row in conn.execute(
                f"""
                SELECT Is_ID, Malzeme, Talep_Donemi, Atanan_Miktar
                FROM "{allocation_table}"
                WHERE Talep_Donemi IS NOT NULL
                  AND Atanan_Miktar IS NOT NULL
                """
            ):
                month = str(row["Talep_Donemi"]).strip()[:7]
                if not re.fullmatch(r"\d{4}-\d{2}", month):
                    continue
                material = normalize_material(row["Malzeme"])
                if not material:
                    continue
                amount = convert_quantity_by_unit(
                    material,
                    float(row["Atanan_Miktar"] or 0.0),
                    unit,
                    gramaj_by_material,
                )
                if amount is None:
                    continue
                months.add(month)
                key = (material, month)
                if row["Is_ID"] is None:
                    stock_served_by_material_month[key] += amount
                else:
                    produced_for_demand_by_material_month[key] += amount

    produced_by_material_month: dict[tuple[str, str], float] = defaultdict(float)
    with sqlite3.connect(plan_db) as conn:
        has_plan_table = (
            conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
                (plan_table,),
            ).fetchone()
            is not None
        )

    if allocation_table_available:
        for key, amount in produced_for_demand_by_material_month.items():
            produced_by_material_month[key] += amount
            months.add(key[1])
    elif has_plan_table:
        plan_query = f"""
            SELECT
                Malzeme,
                Baslangic_Tarihi,
                Bitis_Tarihi,
                Uretim_Gun_Sayisi,
                Ilk_Gun_Uretim_Miktari,
                Diger_Gunler_Gunluk_Uretim_Miktari,
                Toplam_Uretim_Miktari
            FROM "{plan_table}"
            WHERE Baslangic_Tarihi IS NOT NULL
              AND (Uretim_Gun_Sayisi IS NOT NULL OR Toplam_Uretim_Miktari IS NOT NULL)
        """
        with sqlite3.connect(plan_db) as conn:
            conn.row_factory = sqlite3.Row
            for row in conn.execute(plan_query):
                material = normalize_material(row["Malzeme"])
                if not material:
                    continue

                day_count = int(row["Uretim_Gun_Sayisi"] or 0)
                dates = expand_job_dates(row["Baslangic_Tarihi"], row["Bitis_Tarihi"], day_count)
                first_qty = float(row["Ilk_Gun_Uretim_Miktari"] or 0.0)
                other_qty = float(row["Diger_Gunler_Gunluk_Uretim_Miktari"] or 0.0)

                if dates:
                    for idx, work_day in enumerate(dates):
                        month = work_day.strftime("%Y-%m")
                        qty = first_qty if idx == 0 else other_qty
                        amount = convert_quantity_by_unit(material, qty, unit, gramaj_by_material)
                        if amount is None:
                            continue
                        produced_by_material_month[(material, month)] += amount
                        months.add(month)
                    continue

                month = str(row["Baslangic_Tarihi"]).strip()[:7]
                if not re.fullmatch(r"\d{4}-\d{2}", month):
                    continue
                qty = float(row["Toplam_Uretim_Miktari"] or 0.0)
                amount = convert_quantity_by_unit(material, qty, unit, gramaj_by_material)
                if amount is None:
                    continue
                produced_by_material_month[(material, month)] += amount
                months.add(month)
    else:
        for key, amount in fallback_produced_for_demand_by_material_month.items():
            produced_by_material_month[key] += amount
            months.add(key[1])

    all_material_months: set[tuple[str, str]] = set()
    source_maps = (
        fallback_toplam_talep_by_material_month,
        fallback_produced_for_demand_by_material_month,
        produced_for_demand_by_material_month,
        stock_served_by_material_month,
        karsilanan_by_material_month,
        karsilanamayan_by_material_month,
        stok_by_material_month,
        direct_stoktan_by_material_month,
        direct_uretimden_by_material_month,
        produced_by_material_month,
    )
    for source_map in source_maps:
        all_material_months.update(source_map.keys())
    for _material, month in all_material_months:
        months.add(month)

    color_metric_month_values: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: {metric_name: defaultdict(float) for metric_name in USS_METRIC_KEYS}
    )
    for material, month in all_material_months:
        key = (material, month)
        color = material_to_color.get(material, DEFAULT_COLOR_BUCKET)
        if not color:
            color = DEFAULT_COLOR_BUCKET
        metric_bucket = color_metric_month_values[color]

        karsilanan = karsilanan_by_material_month.get(key, 0.0)
        produced_fallback = fallback_produced_for_demand_by_material_month.get(key, 0.0)
        if has_direct_stock_column:
            stoktan_karsilanan = direct_stoktan_by_material_month.get(key, 0.0)
        elif allocation_table_available:
            stoktan_karsilanan = stock_served_by_material_month.get(key, 0.0)
        else:
            stoktan_karsilanan = max(karsilanan - produced_fallback, 0.0)

        if has_direct_production_column:
            uretimden_karsilanan = direct_uretimden_by_material_month.get(key, 0.0)
        elif allocation_table_available:
            uretimden_karsilanan = produced_for_demand_by_material_month.get(key, 0.0)
        else:
            uretimden_karsilanan = max(karsilanan - stoktan_karsilanan, 0.0)

        metric_bucket["Toplam_Talep"][month] += fallback_toplam_talep_by_material_month.get(key, 0.0)
        metric_bucket["Uretilen"][month] += produced_by_material_month.get(key, 0.0)
        metric_bucket["Karsilanan"][month] += karsilanan
        metric_bucket["Stoktan_Karsilanan"][month] += stoktan_karsilanan
        metric_bucket["Uretimden_Karsilanan"][month] += uretimden_karsilanan
        metric_bucket["Karsilanamayan"][month] += karsilanamayan_by_material_month.get(key, 0.0)
        metric_bucket["Stok"][month] += stok_by_material_month.get(key, 0.0)

    months_list = sorted(months)
    used_colors = set(color_metric_month_values.keys())
    ordered_colors = [color for color in color_order if color in used_colors]
    for color in sorted(used_colors - set(ordered_colors), key=sort_key_mixed):
        ordered_colors.append(color)

    return (
        months_list,
        ordered_colors,
        {
            color: {
                metric_name: {
                    month: color_metric_month_values[color][metric_name].get(month, 0.0)
                    for month in months_list
                }
                for metric_name in USS_METRIC_KEYS
            }
            for color in ordered_colors
        },
    )


def monthly_vector(months: list[str]) -> dict[str, float]:
    return {m: 0.0 for m in months}


def safe_ratio(numerator: float, denominator: float) -> float:
    if abs(denominator) < 1e-12:
        return 0.0
    return numerator / denominator


def build_grouped(
    tons: dict[tuple[str, str, str, str], float]
) -> dict[str, dict[str, dict[str, dict[str, float]]]]:
    grouped: dict[str, dict[str, dict[str, dict[str, float]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(dict))
    )
    for (fabrika, firin, hat, month), value in tons.items():
        grouped[fabrika][firin][hat][month] = value
    return grouped


def write_metric_sheet(
    ws,
    months: list[str],
    tons: dict[tuple[str, str, str, str], float],
    structure: dict[str, dict[str, list[str]]],
    source_label: str,
    number_format: str = "0.000",
) -> None:
    columns = ["Kod", *months, "Toplam"]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    grouped = build_grouped(tons)
    all_structure: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for fabrika in structure:
        for firin in structure[fabrika]:
            all_structure[fabrika][firin] = list(structure[fabrika][firin])
    # Any worked line not in Playground_x is still appended.
    for fabrika in grouped:
        for firin in grouped[fabrika]:
            hats = set(all_structure[fabrika][firin])
            hats.update(grouped[fabrika][firin].keys())
            all_structure[fabrika][firin] = sorted(hats, key=sort_key_mixed)

    row_idx = 2
    grand_metric_totals: dict[str, dict[str, float]] = {
        "Uretim": {month: 0.0 for month in months},
        "Stok": {month: 0.0 for month in months},
    }
    grand_production_totals = {month: 0.0 for month in months}
    grand = monthly_vector(months)

    for fabrika in sorted(all_structure, key=sort_key_mixed):
        factory_total = monthly_vector(months)
        firin_totals_for_factory: dict[str, dict[str, float]] = {}
        for firin in sorted(all_structure[fabrika], key=sort_key_mixed):
            firin_total = monthly_vector(months)
            for hat in all_structure[fabrika][firin]:
                month_values = monthly_vector(months)
                for month in months:
                    month_values[month] = grouped[fabrika][firin][hat].get(month, 0.0)
                    firin_total[month] += month_values[month]
                    factory_total[month] += month_values[month]
                    grand[month] += month_values[month]
                row = [f"{fabrika}-{firin}-{hat}", *[month_values[m] for m in months], sum(month_values.values())]
                ws.append(row)
                row_idx += 1

            firin_totals_for_factory[firin] = firin_total

        for firin in sorted(firin_totals_for_factory, key=sort_key_mixed):
            firin_total = firin_totals_for_factory[firin]
            subtotal = [f"Firin Toplam ({fabrika}-{firin})", *[firin_total[m] for m in months], sum(firin_total.values())]
            ws.append(subtotal)
            for col in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col).fill = SUBTOTAL_FILL
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
            row_idx += 1

        factory_row = [f"Fabrika Toplam ({fabrika})", *[factory_total[m] for m in months], sum(factory_total.values())]
        ws.append(factory_row)
        for col in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
        row_idx += 1

    grand_row = ["Genel Toplam", *[grand[m] for m in months], sum(grand.values())]
    ws.append(grand_row)
    for col in range(1, len(columns) + 1):
        ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
        ws.cell(row=row_idx, column=col).font = Font(bold=True)

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
    for col in range(2, len(columns) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = number_format

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 24

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for r in range(2, ws.max_row + 1):
            value = ws.cell(row=r, column=col_idx).value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 11), 60)


def write_ratio_sheet(
    ws,
    months: list[str],
    tons_numerator: dict[tuple[str, str, str, str], float],
    tons_denominator: dict[tuple[str, str, str, str], float],
    structure: dict[str, dict[str, list[str]]],
    source_label: str,
) -> None:
    columns = ["Kod", *months, "Toplam"]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    grouped_num = build_grouped(tons_numerator)
    grouped_den = build_grouped(tons_denominator)
    all_structure: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for fabrika in structure:
        for firin in structure[fabrika]:
            all_structure[fabrika][firin] = list(structure[fabrika][firin])
    for grouped in (grouped_num, grouped_den):
        for fabrika in grouped:
            for firin in grouped[fabrika]:
                hats = set(all_structure[fabrika][firin])
                hats.update(grouped[fabrika][firin].keys())
                all_structure[fabrika][firin] = sorted(hats, key=sort_key_mixed)
    fabrikas = sorted(all_structure.keys(), key=sort_key_mixed)

    row_idx = 2
    grand_num = monthly_vector(months)
    grand_den = monthly_vector(months)

    for fabrika in fabrikas:
        factory_num = monthly_vector(months)
        factory_den = monthly_vector(months)
        firin_totals_for_factory: dict[str, tuple[dict[str, float], dict[str, float]]] = {}
        firins = sorted(all_structure[fabrika].keys(), key=sort_key_mixed)
        for firin in firins:
            firin_num = monthly_vector(months)
            firin_den = monthly_vector(months)
            hats = all_structure[fabrika][firin]
            for hat in hats:
                row_num = monthly_vector(months)
                row_den = monthly_vector(months)
                row_ratio = monthly_vector(months)
                for month in months:
                    n_val = grouped_num[fabrika][firin][hat].get(month, 0.0)
                    d_val = grouped_den[fabrika][firin][hat].get(month, 0.0)
                    row_num[month] = n_val
                    row_den[month] = d_val
                    row_ratio[month] = safe_ratio(n_val, d_val)
                    firin_num[month] += n_val
                    firin_den[month] += d_val
                    factory_num[month] += n_val
                    factory_den[month] += d_val
                    grand_num[month] += n_val
                    grand_den[month] += d_val

                total_ratio = safe_ratio(sum(row_num.values()), sum(row_den.values()))
                row = [f"{fabrika}-{firin}-{hat}", *[row_ratio[m] for m in months], total_ratio]
                ws.append(row)
                row_idx += 1

            firin_totals_for_factory[firin] = (firin_num, firin_den)

        for firin in firins:
            firin_num, firin_den = firin_totals_for_factory[firin]
            subtotal = [
                f"Firin Toplam ({fabrika}-{firin})",
                *[safe_ratio(firin_num[m], firin_den[m]) for m in months],
                safe_ratio(sum(firin_num.values()), sum(firin_den.values())),
            ]
            ws.append(subtotal)
            for col in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col).fill = SUBTOTAL_FILL
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
            row_idx += 1

        factory_row = [
            f"Fabrika Toplam ({fabrika})",
            *[safe_ratio(factory_num[m], factory_den[m]) for m in months],
            safe_ratio(sum(factory_num.values()), sum(factory_den.values())),
        ]
        ws.append(factory_row)
        for col in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
        row_idx += 1

    grand_row = [
        "Genel Toplam",
        *[safe_ratio(grand_num[m], grand_den[m]) for m in months],
        safe_ratio(sum(grand_num.values()), sum(grand_den.values())),
    ]
    ws.append(grand_row)
    for col in range(1, len(columns) + 1):
        ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
        ws.cell(row=row_idx, column=col).font = Font(bold=True)

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
    for col in range(2, len(columns) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "0.0000"

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 24

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for r in range(2, ws.max_row + 1):
            value = ws.cell(row=r, column=col_idx).value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 11), 60)


def write_factory_color_summary_sheet(
    ws,
    months: list[str],
    color_rows: list[str],
    prod_by_factory_color_month: dict[tuple[str, str, str], float],
    stock_by_factory_color_month: dict[tuple[str, str, str], float],
    served_by_factory_color_month: dict[tuple[str, str, str], float],
    structure: dict[str, dict[str, list[str]]],
    source_label: str,
) -> None:
    columns = ["Fabrika", "Tip", "Renk", *months, "Toplam"]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    factories = sorted(structure.keys(), key=sort_key_mixed)
    # Keep any factory that appears in metrics but not in structure.
    metrics_factories = set()
    for f, _c, _m in prod_by_factory_color_month:
        metrics_factories.add(f)
    for f, _c, _m in stock_by_factory_color_month:
        metrics_factories.add(f)
    for f, _c, _m in served_by_factory_color_month:
        metrics_factories.add(f)
    for factory in sorted(metrics_factories - set(factories), key=sort_key_mixed):
        factories.append(factory)

    row_idx = 2
    grand_metric_totals: dict[str, dict[str, float]] = {
        "Uretim": {month: 0.0 for month in months},
        "Stok": {month: 0.0 for month in months},
    }

    def metric_value(metric_name: str, factory: str, color: str, month: str) -> float:
        if metric_name == "Uretim":
            return prod_by_factory_color_month.get((factory, color, month), 0.0)
        if metric_name == "Stok":
            return stock_by_factory_color_month.get((factory, color, month), 0.0)
        if metric_name == "Karsilanan":
            return served_by_factory_color_month.get((factory, color, month), 0.0)
        return 0.0

    metric_specs = [
        ("Uretim", "Toplam Uretim"),
        ("Stok", "Toplam Stok"),
        ("Karsilanan", "Toplam Karsilanan"),
    ]

    for factory in factories:
        for metric_name, total_label in metric_specs:
            totals = {month: 0.0 for month in months}
            for color in color_rows:
                vals = [metric_value(metric_name, factory, color, month) for month in months]
                for month, value in zip(months, vals):
                    totals[month] += value
                ws.append([factory, metric_name, color, *vals, sum(vals)])
                row_idx += 1

            total_vals = [totals[month] for month in months]
            ws.append([factory, metric_name, total_label, *total_vals, sum(total_vals)])
            for col in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col).fill = SUBTOTAL_FILL
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
            if metric_name in grand_metric_totals:
                for month, value in zip(months, total_vals):
                    grand_metric_totals[metric_name][month] += value
            row_idx += 1

        ws.append(["", "", "", *["" for _ in months], ""])
        row_idx += 1

    # Final control rows for cross-sheet totals.
    grand_rows = [
        ("Uretim", "Toplam Uretim"),
        ("Stok", "Toplam Stok"),
    ]
    for metric_name, total_label in grand_rows:
        month_totals = [grand_metric_totals[metric_name][month] for month in months]
        ws.append(["Genel Toplam", metric_name, total_label, *month_totals, sum(month_totals)])
        for col in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
        row_idx += 1

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    for col in range(4, len(columns) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "0.000"

    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 24

    widths = {1: 12, 2: 14, 3: 36}
    for col_idx, col_name in enumerate(columns, start=1):
        if col_idx in widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx]
            continue
        max_len = len(str(col_name))
        for r in range(2, ws.max_row + 1):
            text = str(ws.cell(row=r, column=col_idx).value or "")
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 16)


def write_uss_summary_sheet(
    ws,
    months: list[str],
    metric_month_values: dict[str, dict[str, float]],
    source_label: str,
    number_format: str = "0",
) -> None:
    columns = ["Tip", *months, "Toplam"]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for metric_key, label in USS_METRIC_ROWS:
        values = [metric_month_values.get(metric_key, {}).get(month, 0.0) for month in months]
        ws.append([label, *values, sum(values)])

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    for col in range(2, len(columns) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = number_format

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 24

    widths = {1: 24}
    for col_idx, col_name in enumerate(columns, start=1):
        if col_idx in widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx]
            continue
        max_len = len(str(col_name))
        for r in range(2, ws.max_row + 1):
            text = str(ws.cell(row=r, column=col_idx).value or "")
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 16)


def write_uss_color_summary_sheet(
    ws,
    months: list[str],
    color_rows: list[str],
    color_metric_month_values: dict[str, dict[str, dict[str, float]]],
    source_label: str,
    number_format: str = "0",
) -> None:
    columns = ["Renk", "Tip", *months, "Toplam"]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for color in color_rows:
        metric_map = color_metric_month_values.get(color, {})
        for metric_key, label in USS_METRIC_ROWS:
            values = [metric_map.get(metric_key, {}).get(month, 0.0) for month in months]
            ws.append([color, label, *values, sum(values)])

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    for col in range(3, len(columns) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = number_format

    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 24

    widths = {1: 36, 2: 30}
    for col_idx, col_name in enumerate(columns, start=1):
        if col_idx in widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx]
            continue
        max_len = len(str(col_name))
        for r in range(2, ws.max_row + 1):
            text = str(ws.cell(row=r, column=col_idx).value or "")
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 16)


def period_text_from_month_start(month_start: date) -> str:
    return f"{month_start.day}.{month_start.month:02d}.{month_start.year}"


def collect_upload_production_plan_rows(
    plan_db: Path,
    plan_table: str,
) -> list[tuple[str, str, str, str, int, str, str]]:
    monthly_qty: dict[tuple[str, str, date], float] = defaultdict(float)
    query = f"""
        SELECT
            Malzeme,
            Fabrika,
            Baslangic_Tarihi,
            Bitis_Tarihi,
            Uretim_Gun_Sayisi,
            Ilk_Gun_Uretim_Miktari,
            Diger_Gunler_Gunluk_Uretim_Miktari
        FROM "{plan_table}"
        WHERE Baslangic_Tarihi IS NOT NULL
          AND Bitis_Tarihi IS NOT NULL
          AND Uretim_Gun_Sayisi IS NOT NULL
    """

    with sqlite3.connect(plan_db) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(query):
            item_code = normalize_material(row["Malzeme"])
            location_code = str(row["Fabrika"]).strip()
            if not item_code or not location_code:
                continue

            day_count = int(row["Uretim_Gun_Sayisi"] or 0)
            dates = expand_job_dates(row["Baslangic_Tarihi"], row["Bitis_Tarihi"], day_count)
            if not dates:
                continue

            first_qty = float(row["Ilk_Gun_Uretim_Miktari"] or 0.0)
            other_qty = float(row["Diger_Gunler_Gunluk_Uretim_Miktari"] or 0.0)

            for idx, work_day in enumerate(dates):
                qty = first_qty if idx == 0 else other_qty
                month_start = work_day.replace(day=1)
                monthly_qty[(item_code, location_code, month_start)] += qty

    rows = []
    for (item_code, location_code, month_start), qty in sorted(
        monthly_qty.items(), key=lambda x: (x[0][2], sort_key_mixed(x[0][1]), sort_key_mixed(x[0][0]))
    ):
        rows.append(
            (
                item_code,
                location_code,
                period_text_from_month_start(month_start),
                "",
                int(round(qty)),
                "Production",
                "ADT",
            )
        )
    return rows


def write_upload_production_plan_sheet(
    ws,
    rows: list[tuple[str, str, str, str, int, str, str]],
    source_label: str,
) -> None:
    columns = [
        "ItemCode",
        "LocationCode",
        "Period",
        "SupplyCategory",
        "SupplyQuantity",
        "SupplyType",
        "UnitOfMeasure",
    ]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in rows:
        ws.append(list(row))

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "0"

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    widths = {1: 14, 2: 14, 3: 12, 4: 16, 5: 16, 6: 12, 7: 14}
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 14)


def write_report(
    output_xlsx: Path,
    months_gramaj: list[str],
    tons_gramaj: dict[tuple[str, str, str, str], float],
    months_gramaj_adet: list[str],
    adet_gramaj: dict[tuple[str, str, str, str], float],
    missing: dict[str, MissingGramaj],
    months_cekis: list[str],
    tons_cekis: dict[tuple[str, str, str, str], float],
    months_changes: list[str],
    change_counts: dict[tuple[str, str, str, str], float],
    months_color_summary: list[str],
    color_rows: list[str],
    prod_by_factory_color_month: dict[tuple[str, str, str], float],
    stock_by_factory_color_month: dict[tuple[str, str, str], float],
    served_by_factory_color_month: dict[tuple[str, str, str], float],
    months_color_summary_adet: list[str],
    color_rows_adet: list[str],
    prod_by_factory_color_month_adet: dict[tuple[str, str, str], float],
    stock_by_factory_color_month_adet: dict[tuple[str, str, str], float],
    served_by_factory_color_month_adet: dict[tuple[str, str, str], float],
    months_uss: list[str],
    uss_metric_month_values: dict[str, dict[str, float]],
    months_uss_adet: list[str],
    uss_metric_month_values_adet: dict[str, dict[str, float]],
    months_uss_color: list[str],
    uss_color_rows: list[str],
    uss_color_metric_month_values: dict[str, dict[str, dict[str, float]]],
    months_uss_color_adet: list[str],
    uss_color_rows_adet: list[str],
    uss_color_metric_month_values_adet: dict[str, dict[str, dict[str, float]]],
    upload_rows: list[tuple[str, str, str, str, int, str, str]],
    structure: dict[str, dict[str, list[str]]],
    plan_table: str,
) -> None:
    wb = Workbook()

    ws_gramaj = wb.active
    ws_gramaj.title = "Aylik_Uretim"
    write_metric_sheet(
        ws_gramaj,
        months_gramaj,
        tons_gramaj,
        structure,
        f"Kod (Kaynak: {plan_table} | Metric: Toplam_Uretim_Miktari x Gramaj / 1e6 | Gramaj=Fiili oncelikli, fallback=Net/Talepler)",
    )

    ws_gramaj_adet = wb.create_sheet("Aylik_Uretim_Adet")
    write_metric_sheet(
        ws_gramaj_adet,
        months_gramaj_adet,
        adet_gramaj,
        structure,
        f"Kod (Kaynak: {plan_table} | Metric: Toplam_Uretim_Miktari)",
        number_format="0",
    )

    ws_cekis = wb.create_sheet("Aylik_Uretim_Firin_Cekisi")
    write_metric_sheet(
        ws_cekis,
        months_cekis,
        tons_cekis,
        structure,
        f"Kod (Kaynak: {plan_table} | Metric: Toplam_Firin_Cekisi)",
    )

    months_ratio = sorted(set(months_gramaj) | set(months_cekis))
    ws_ratio = wb.create_sheet("Aylik_Uretim_Oran")
    write_ratio_sheet(
        ws_ratio,
        months_ratio,
        tons_gramaj,
        tons_cekis,
        structure,
        f"Kod (Kaynak: {plan_table} | Metric: (Toplam_Uretim_Miktari x Gramaj / 1e6) / Toplam_Firin_Cekisi | Gramaj=Fiili oncelikli, fallback=Net/Talepler)",
    )

    ws_change = wb.create_sheet("Aylik_Uretim_Degisim_Sayisi")
    write_metric_sheet(
        ws_change,
        months_changes,
        change_counts,
        structure,
        f"Kod (Kaynak: {plan_table} | Metric: Uretim Degisim Sayisi)",
        number_format="0",
    )

    ws_color_summary = wb.create_sheet("Renk_Bazli_Fabrika_Ozet")
    write_factory_color_summary_sheet(
        ws_color_summary,
        months_color_summary,
        color_rows,
        prod_by_factory_color_month,
        stock_by_factory_color_month,
        served_by_factory_color_month,
        structure,
        f"Kaynak: {plan_table} | Uretim=Qty*Gramaj/1e6 (Fiili oncelikli, fallback=Net/Talepler), Stok=Donem Sonu, Karsilanan=Atanan",
    )

    ws_color_summary_adet = wb.create_sheet("Renk_Bazli_Fabrika_Ozet_Adet")
    write_factory_color_summary_sheet(
        ws_color_summary_adet,
        months_color_summary_adet,
        color_rows_adet,
        prod_by_factory_color_month_adet,
        stock_by_factory_color_month_adet,
        served_by_factory_color_month_adet,
        structure,
        f"Kaynak: {plan_table} | Uretim=Qty, Stok=Donem Sonu, Karsilanan=Atanan",
    )

    ws_uss = wb.create_sheet("USS")
    write_uss_summary_sheet(
        ws_uss,
        months_uss,
        uss_metric_month_values,
        f"Kaynak: Talepler + Full_ORTools_Plan_Talep_Stok + {plan_table} | Toplam_Talep=Talepler.[MM.YYYY ADT]*1000 (ton=ADT*1000*Fiili_Gramaj_Tekil/1e6, fallback=Net_Agirligi), Uretilen=talebe atanan uretim, Karsilanan=Karsilanan_Talep_Miktari, Stoktan_Karsilanan=Stoktan_Karsilanan_Miktar, Uretimden_Karsilanan=Uretimden_Karsilanan_Miktar, Karsilanamayan=Kalan_Acik_Talep, Stok=Donem_Sonu_Stok",
    )

    ws_uss_adet = wb.create_sheet("USS_Adet")
    write_uss_summary_sheet(
        ws_uss_adet,
        months_uss_adet,
        uss_metric_month_values_adet,
        f"Kaynak: Talepler + Full_ORTools_Plan_Talep_Stok + {plan_table} | Toplam_Talep=Talepler.[MM.YYYY ADT]*1000, Uretilen=talebe atanan uretim, Karsilanan=Karsilanan_Talep_Miktari, Stoktan_Karsilanan=Stoktan_Karsilanan_Miktar, Uretimden_Karsilanan=Uretimden_Karsilanan_Miktar, Karsilanamayan=Kalan_Acik_Talep, Stok=Donem_Sonu_Stok",
        number_format="0",
    )

    ws_uss_color = wb.create_sheet("USS_Renk_Bazli_Ozet")
    write_uss_color_summary_sheet(
        ws_uss_color,
        months_uss_color,
        uss_color_rows,
        uss_color_metric_month_values,
        f"Kaynak: Full_ORTools_Plan_Talep_Stok + Full_ORTools_Plan_Is_Talep_Esleme + {plan_table} + talepler_x | Renk=Malzeme->Renk eslesmesi",
    )

    ws_uss_color_adet = wb.create_sheet("USS_Renk_Bazli_Ozet_Adet")
    write_uss_color_summary_sheet(
        ws_uss_color_adet,
        months_uss_color_adet,
        uss_color_rows_adet,
        uss_color_metric_month_values_adet,
        f"Kaynak: Full_ORTools_Plan_Talep_Stok + Full_ORTools_Plan_Is_Talep_Esleme + {plan_table} + talepler_x | Renk=Malzeme->Renk eslesmesi",
        number_format="0",
    )

    ws_upload = wb.create_sheet("Yuklenecek Uretim Plani")
    write_upload_production_plan_sheet(
        ws_upload,
        upload_rows,
        f"Kaynak: {plan_table} | Format: Upload Plan",
    )
    wb.save(output_xlsx)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="SQLite verisinden Fabrika-Firin-Hat aylik uretim ton Excel raporu uretir.")
    parser.add_argument("--plan-db", default="planlama.sqlite", help="Planlama SQLite yolu")
    parser.add_argument("--etl-db", default="etl.sqlite", help="ETL SQLite yolu")
    parser.add_argument("--plan-table", default="Full_ORTools_Plan_Uretim_Isleri", help="Kaynak plan tablosu")
    parser.add_argument("--output", default="aylik_uretim_fabrika_firin_hat.xlsx", help="Cikti Excel yolu")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    plan_db = Path(args.plan_db)
    etl_db = Path(args.etl_db)
    output = Path(args.output)

    emit_progress(5, "load_structure", "Hat yapisi okunuyor")
    structure = load_playground_structure(plan_db)
    emit_progress(18, "load_gramaj", "Gramaj verisi okunuyor")
    gramaj_by_material = resolve_production_ton_gramaj_map(etl_db)
    emit_progress(30, "metric_monthly_tons", "Aylik uretim metrikleri hesaplaniyor")
    tons_gramaj, months_gramaj, missing = collect_monthly_tons_by_gramaj(plan_db, args.plan_table, gramaj_by_material)
    adet_gramaj, months_gramaj_adet = collect_monthly_adet_by_production(plan_db, args.plan_table)
    emit_progress(44, "metric_furnace_draw", "Firin cekisi metrikleri hesaplaniyor")
    tons_cekis, months_cekis = collect_monthly_tons_by_firin_cekisi(plan_db, args.plan_table)
    emit_progress(56, "metric_change_count", "Aylik degisim sayilari hesaplaniyor")
    change_counts, months_changes = collect_monthly_change_counts(plan_db, args.plan_table)
    emit_progress(68, "metric_color_summary", "Renk bazli ozet metrikleri hazirlaniyor")
    (
        months_color_summary,
        color_rows,
        prod_by_factory_color_month,
        stock_by_factory_color_month,
        served_by_factory_color_month,
        color_diag,
    ) = collect_color_factory_month_metrics(plan_db, etl_db, args.plan_table, unit="ton")
    (
        months_color_summary_adet,
        color_rows_adet,
        prod_by_factory_color_month_adet,
        stock_by_factory_color_month_adet,
        served_by_factory_color_month_adet,
        _color_diag_adet,
    ) = collect_color_factory_month_metrics(plan_db, etl_db, args.plan_table, unit="adet")
    emit_progress(74, "metric_uss", "USS metrikleri hazirlaniyor")
    months_uss, uss_metric_month_values = collect_uss_month_metrics(
        plan_db, etl_db, unit="ton", plan_table=args.plan_table
    )
    months_uss_adet, uss_metric_month_values_adet = collect_uss_month_metrics(
        plan_db, etl_db, unit="adet", plan_table=args.plan_table
    )
    (
        months_uss_color,
        uss_color_rows,
        uss_color_metric_month_values,
    ) = collect_uss_color_month_metrics(
        plan_db,
        etl_db,
        unit="ton",
        plan_table=args.plan_table,
    )
    (
        months_uss_color_adet,
        uss_color_rows_adet,
        uss_color_metric_month_values_adet,
    ) = collect_uss_color_month_metrics(
        plan_db,
        etl_db,
        unit="adet",
        plan_table=args.plan_table,
    )
    emit_progress(82, "upload_rows", "Yukleme plani satirlari uretiliyor")
    upload_rows = collect_upload_production_plan_rows(plan_db, args.plan_table)

    if (
        not months_gramaj
        and not months_gramaj_adet
        and not months_cekis
        and not months_changes
        and not months_color_summary
        and not months_color_summary_adet
        and not months_uss
        and not months_uss_adet
        and not months_uss_color
        and not months_uss_color_adet
        and not upload_rows
    ):
        raise SystemExit("Rapor uretilemedi: kaynak tabloda gecerli ay bulunamadi.")

    emit_progress(90, "write_report", "Rapor dosyasi yaziliyor")
    write_report(
        output,
        months_gramaj,
        tons_gramaj,
        months_gramaj_adet,
        adet_gramaj,
        missing,
        months_cekis,
        tons_cekis,
        months_changes,
        change_counts,
        months_color_summary,
        color_rows,
        prod_by_factory_color_month,
        stock_by_factory_color_month,
        served_by_factory_color_month,
        months_color_summary_adet,
        color_rows_adet,
        prod_by_factory_color_month_adet,
        stock_by_factory_color_month_adet,
        served_by_factory_color_month_adet,
        months_uss,
        uss_metric_month_values,
        months_uss_adet,
        uss_metric_month_values_adet,
        months_uss_color,
        uss_color_rows,
        uss_color_metric_month_values,
        months_uss_color_adet,
        uss_color_rows_adet,
        uss_color_metric_month_values_adet,
        upload_rows,
        structure,
        args.plan_table,
    )

    emit_progress(98, "finalize", "Rapor ozeti tamamlaniyor")
    print(f"Rapor olusturuldu: {output.resolve()}")
    print(
        f"Gramaj sheet ay sayisi: {len(months_gramaj)} | Detay kayit: {len(tons_gramaj)} | "
        f"Eksik gramaj malzeme: {len(missing)}"
    )
    print(f"Firin cekisi sheet ay sayisi: {len(months_cekis)} | Detay kayit: {len(tons_cekis)}")
    print(f"Degisim sayisi sheet ay sayisi: {len(months_changes)} | Detay kayit: {len(change_counts)}")
    print(
        "Renk ozet: "
        f"ay={len(months_color_summary)} renk={len(color_rows)} "
        f"null_ton={color_diag['null_alloc_total_ton']:.3f} "
        f"dagitilan={color_diag['null_alloc_distributed_ton']:.3f} "
        f"dagitilmayan={color_diag['null_alloc_undistributed_ton']:.3f} "
        f"fixed_kisa={color_diag['fixed_alloc_shortfall_ton']:.3f}"
    )
    uss_ton_total = sum(sum(values.values()) for values in uss_metric_month_values.values())
    uss_adet_total = sum(sum(values.values()) for values in uss_metric_month_values_adet.values())
    print(f"USS sheet ay sayisi (ton/adet): {len(months_uss)}/{len(months_uss_adet)} | Toplam (ton): {uss_ton_total:.3f} | Toplam (adet): {uss_adet_total:.0f}")
    uss_color_ton_total = sum(
        sum(sum(month_values.values()) for month_values in metric_map.values())
        for metric_map in uss_color_metric_month_values.values()
    )
    uss_color_adet_total = sum(
        sum(sum(month_values.values()) for month_values in metric_map.values())
        for metric_map in uss_color_metric_month_values_adet.values()
    )
    print(
        "USS renk ozet sheet ay sayisi (ton/adet): "
        f"{len(months_uss_color)}/{len(months_uss_color_adet)} "
        f"| Renk sayisi (ton/adet): {len(uss_color_rows)}/{len(uss_color_rows_adet)} "
        f"| Toplam (ton): {uss_color_ton_total:.3f} "
        f"| Toplam (adet): {uss_color_adet_total:.0f}"
    )
    print(f"Yuklenecek uretim plani satir sayisi: {len(upload_rows)}")
    emit_progress(100, "completed", "Aylik rapor adimi tamamlandi")


if __name__ == "__main__":
    main()
